# -*- coding: utf-8 -*-
"""
发票明细导出Excel工具（V4.1 - 最终完整版：默认保存路径改为当前登录用户桌面）
"""

import os
import re
import traceback
import configparser
import threading
import subprocess  # 用于非Windows系统打开文件
from datetime import datetime
from tkinter import (
    Tk, Frame, Button, Label, StringVar, Radiobutton, messagebox, filedialog, END, Toplevel,
    DoubleVar
)
from tkinter import ttk
import pdfplumber
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

CONFIG_FILE = 'invoice_config.ini'
UNIT_KEYWORDS = [
    '千克', '个', '件', '套', '台', '张', '米', '公斤', '升', '吨',
    '箱', '盒', '包', '瓶', '罐', '条', '只', '卷', '桶', '大', '袋',
    '块', '次', '批', '批次', '项',
    '枚', '支', '根', '头', '辆', '架', '艘', '本', '册', '部',
    '组', '副', '双', '对', '厘米', '毫米', '公里', '英尺', '英寸',
    '平方米', '平方英尺', '立方米', '立方厘米', '加仑', '毫升',
    '克', '毫克', '磅', '盎司', '小时', '天', '月', '年', '季度',
    '份', '页', '场', '位', '人次', '立方', '平米', '度', '千瓦时', '斤', '两',
    '㎡', 'm³', '㎏', '㎞', '千米', '公里', '克'
]

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Invoice2EXCEL V4.2")
        self.config = configparser.ConfigParser()
        self.load_config()

        # 固定窗口大小为 1139x837
        fixed_width = 1139
        fixed_height = 837
        self.root.geometry(f"{fixed_width}x{fixed_height}")

        # 禁止调整窗口大小
        self.root.resizable(False, False)

        # 恢复上次窗口位置
        try:
            geom = self.config.get('window', 'position', fallback=None)
            if geom:
                if '+' in geom:
                    pos_part = geom
                    if '+' in pos_part[1:]:
                        x, y = pos_part.split('+', 2)[1:]
                        if x.lstrip('-').isdigit() and y.lstrip('-').isdigit():
                            self.root.geometry(f"+{x}+{y}")
        except:
            pass

        # 修改默认保存路径为当前登录用户的桌面（兼容所有用户）
        self.save_path = self.config.get('settings', 'save_path', fallback=os.path.join(os.path.expanduser("~"), "Desktop"))

        self.invoices = []
        self.export_mode = StringVar(value=self.config.get('settings', 'export_mode', fallback="merge"))
        self.all_var = StringVar(value=self.config.get('settings', 'all_selected', fallback="1"))
        self.loading_dialog = None
        self.progress_var = None
        self.progress_label = None
        self.build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.tree.bind('<Button-1>', self.on_check_click)
        self.tree.tag_configure('success', background='#e6f3ff')
        self.root.update_idletasks()
        self.root.after(100, self.force_column_widths)

    def force_column_widths(self):
        self.tree.column('check', width=60)
        self.tree.column('num', width=80)
        self.tree.column('file', width=800)
        self.tree.column('rows', width=150)
        self.root.update_idletasks()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                self.config.read(CONFIG_FILE, encoding='utf-8')
            except:
                pass

    def save_config(self):
        try:
            if 'settings' not in self.config:
                self.config.add_section('settings')
            if 'window' not in self.config:
                self.config.add_section('window')

            self.config['settings']['save_path'] = self.save_path
            self.config['settings']['export_mode'] = self.export_mode.get()
            self.config['settings']['all_selected'] = self.all_var.get()

            # 保存窗口位置
            current_geom = self.root.geometry()
            if '+' in current_geom:
                pos_part = '+' + current_geom.split('+', 1)[1]
                self.config['window']['position'] = pos_part

            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                self.config.write(f)
        except Exception as e:
            print(f"保存配置失败: {e}")

    def on_closing(self):
        self.save_config()
        self.root.destroy()

    def build_ui(self):
        top_frame = Frame(self.root, bg='#f0f0f0', pady=10)
        top_frame.pack(fill='x', padx=10, pady=5)

        btn_width = 12
        Button(top_frame, text="添加发票", command=self.add_files, width=btn_width, height=1,
               font=('微软雅黑', 10), bg='#4a90e2', fg='white').pack(side='left', padx=5)
        Button(top_frame, text="添加文件夹", command=self.add_folder, width=btn_width, height=1,
               font=('微软雅黑', 10), bg='#4a90e2', fg='white').pack(side='left', padx=5)
        Label(top_frame, text="支持多选、多页电子发票PDF文件", font=('微软雅黑', 10), bg='#f0f0f0').pack(side='left', padx=10)

        mid_frame = Frame(self.root)
        mid_frame.pack(fill='both', expand=True, padx=10, pady=(0, 5))

        header_frame = Frame(mid_frame)
        header_frame.pack(fill='x')
        Label(header_frame, text="已添加的发票列表",
              font=('微软雅黑', 9)).pack(anchor='w', padx=5)

        tree_container = Frame(mid_frame)
        tree_container.pack(fill='both', expand=True)

        columns = ('check', 'num', 'file', 'rows')
        self.tree = ttk.Treeview(tree_container, columns=columns, show='headings', selectmode='extended')

        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.heading('check', text='✓')
        self.tree.heading('num', text='序号')
        self.tree.heading('file', text='发票文件')
        self.tree.heading('rows', text='明细行数')

        self.tree.column('check', width=60, anchor='center')
        self.tree.column('num', width=80, anchor='center')
        self.tree.column('file', width=800, anchor='w')
        self.tree.column('rows', width=150, anchor='center')

        self.tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview.Heading',
                        background='#e0e0e0',
                        foreground='black',
                        relief='raised',
                        font=('微软雅黑', 10, 'bold'),
                        padding=8)
        style.map('Treeview.Heading',
                  background=[('active', '#d0d0d0')],
                  relief=[('pressed', '!disabled', 'sunken')])

        self.selected_count_frame = Frame(self.root)
        self.selected_count_frame.pack(fill='x', padx=10, pady=(0, 5))
        self.selected_count_label = Label(self.selected_count_frame, text="已有 0 条符合条件的发票文件被选中",
                                          font=('微软雅黑', 9), fg='gray', anchor='w')
        self.selected_count_label.pack(side='left')

        ctrl_frame = Frame(self.root)
        ctrl_frame.pack(fill='x', padx=10, pady=5)
        ttk.Checkbutton(ctrl_frame, text="全选", variable=self.all_var,
                        command=self.toggle_all).pack(side='left', padx=5)
        Button(ctrl_frame, text="反选", command=self.invert_selection,
               width=8, bg='#f39c12', fg='white').pack(side='left', padx=5)
        Button(ctrl_frame, text="删除选中", command=self.delete_selected,
               width=12, bg='#e74c3c', fg='white').pack(side='left', padx=10)

        bottom_frame = Frame(self.root, pady=10)
        bottom_frame.pack(fill='x', padx=10)

        left_frame = Frame(bottom_frame)
        left_frame.pack(side='left', fill='y', padx=10)
        Label(left_frame, text="导出选项:", font=('微软雅黑', 10, 'bold')).pack(anchor='w')
        mode_frame = Frame(left_frame)
        mode_frame.pack(anchor='w', pady=5)
        Radiobutton(mode_frame, text="合并导出（所有发票合并到一个SHEET）",
                    variable=self.export_mode, value="merge",
                    font=('微软雅黑', 9)).pack(anchor='w')
        Radiobutton(mode_frame, text="分表导出（每张发票单独一个SHEET）",
                    variable=self.export_mode, value="separate",
                    font=('微软雅黑', 9)).pack(anchor='w', pady=2)

        middle_frame = Frame(bottom_frame)
        middle_frame.pack(side='left', fill='both', expand=True, padx=20)
        Label(middle_frame, text="保存路径:", font=('微软雅黑', 10, 'bold')).pack(anchor='w')
        path_btn_frame = Frame(middle_frame)
        path_btn_frame.pack(anchor='w', pady=5)
        Button(path_btn_frame, text="浏览...", command=self.set_save_path,
               width=10).pack(side='left')
        Button(path_btn_frame, text="打开文件夹", command=self.open_save_folder,
               width=10).pack(side='left', padx=5)
        path_text = self.shorten_path(self.save_path)
        self.path_label = Label(middle_frame, text=path_text,
                                fg='blue', font=('微软雅黑', 9),
                                wraplength=400, justify='left', anchor='w')
        self.path_label.pack(anchor='w', pady=2)

        right_frame = Frame(bottom_frame)
        right_frame.pack(side='right', padx=10)
        Button(right_frame, text="导出EXCEL", command=self.export_to_excel,
               width=btn_width, height=1, font=('微软雅黑', 10),
               bg='#27ae60', fg='white').pack()

        current_year = datetime.now().year
        copyright_text = f"Developed by Nero Diao ©2026-{current_year}"
        copyright_label = Label(self.root, text=copyright_text,
                                font=('微软雅黑', 8), fg='#888888', anchor='w')
        copyright_label.pack(side='bottom', fill='x', padx=15, pady=(0, 8))

    def shorten_path(self, p):
        p = os.path.normpath(p)
        if len(p) <= 50:
            return p
        return "..." + p[-47:]

    def invert_selection(self):
        for item in self.tree.get_children():
            values = list(self.tree.item(item, 'values'))
            values[0] = '☐' if values[0] == '✓' else '✓'
            self.tree.item(item, values=values)
        self.update_selected_count()

    def on_check_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        if not item or col != '#1':
            return
        values = list(self.tree.item(item, 'values'))
        values[0] = '✓' if values[0] == '☐' else '☐'
        self.tree.item(item, values=values)
        self.update_all_var()
        self.update_selected_count()

    def toggle_all(self):
        mark = '✓' if self.all_var.get() == "1" else '☐'
        for item in self.tree.get_children():
            values = list(self.tree.item(item, 'values'))
            values[0] = mark
            self.tree.item(item, values=values)
        self.update_selected_count()

    def update_all_var(self):
        checks = [self.tree.item(item, 'values')[0] for item in self.tree.get_children()]
        if not checks:
            self.all_var.set("0")
        elif all(c == '✓' for c in checks):
            self.all_var.set("1")
        else:
            self.all_var.set("0")

    def update_selected_count(self):
        selected_count = sum(1 for item in self.tree.get_children() if self.tree.item(item, 'values')[0] == '✓')
        self.selected_count_label.config(text=f"已有 {selected_count} 条符合条件的发票文件被选中")

    def create_progress_dialog(self, total_files):
        self.loading_dialog = Toplevel(self.root)
        self.loading_dialog.title("发票导入中")
        self.loading_dialog.transient(self.root)
        self.loading_dialog.grab_set()
        self.loading_dialog.resizable(False, False)
        self.loading_dialog.configure(bg='#f8f9fa')
        dialog_width = 420
        dialog_height = 180
        self.root.update_idletasks()
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        pos_x = root_x + (root_width - dialog_width) // 2
        pos_y = root_y + (root_height - dialog_height) // 2
        self.loading_dialog.geometry(f"{dialog_width}x{dialog_height}+{pos_x}+{pos_y}")
        main_frame = Frame(self.loading_dialog, bg='#f8f9fa', padx=30, pady=30)
        main_frame.pack(fill='both', expand=True)
        Label(main_frame, text="正在导入并解析发票文件，请稍候...",
              font=('微软雅黑', 11, 'bold'), bg='#f8f9fa', fg='#333333').pack(pady=(0, 20))
        self.progress_label = Label(main_frame, text=f"当前正在处理第 1 条 / 共 {total_files} 条",
                                    font=('微软雅黑', 12), bg='#f8f9fa', fg='#2c3e50')
        self.progress_label.pack(pady=(0, 20))
        self.progress_var = DoubleVar()
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Blue.Horizontal.TProgressbar",
                        background='#4a90e2',
                        troughcolor='#e0e0e0',
                        borderwidth=0,
                        lightcolor='#4a90e2',
                        darkcolor='#4a90e2',
                        thickness=24)
        progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=total_files,
                                       style="Blue.Horizontal.TProgressbar", length=360)
        progress_bar.pack(pady=(0, 10))

    def update_progress(self, current, total):
        if self.progress_var is not None:
            self.progress_var.set(current)
        if self.progress_label is not None:
            self.progress_label.config(text=f"当前正在处理第 {current} 条 / 共 {total} 条")

    def safe_close_loading_dialog(self):
        if self.loading_dialog and self.loading_dialog.winfo_exists():
            self.loading_dialog.destroy()
        self.loading_dialog = None
        self.progress_var = None
        self.progress_label = None

    def has_invoice_number_in_first_two_pages(self, pdf_path):
        invoice_no_patterns = [
            r'发票号码[：:\s]*([0-9]{8,12})',
            r'发票代码[：:\s]*([0-9]{10,12})',
            r'No[.:：\s]*([0-9]{8,12})',
            r'发票号[：:\s]*([0-9]{8,12})',
        ]
        try:
            with pdfplumber.open(pdf_path) as pdf:
                pages_to_check = pdf.pages[:2]
                for page in pages_to_check:
                    text = page.extract_text()
                    if not text:
                        continue
                    text = text.upper()
                    for pattern in invoice_no_patterns:
                        if re.search(pattern, text):
                            return True
                return False
        except:
            return False

    def add_folder(self):
        folder = filedialog.askdirectory(title="选择包含发票PDF的文件夹")
        if not folder:
            return
        pdf_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith('.pdf')]
        if not pdf_files:
            messagebox.showinfo("提示", "所选文件夹中未找到PDF文件")
            return
        self.create_progress_dialog(len(pdf_files))
        def import_thread():
            self.process_files(pdf_files)
            self.root.after(0, self.safe_close_loading_dialog)
        threading.Thread(target=import_thread, daemon=True).start()

    def add_files(self):
        files = filedialog.askopenfilenames(
            title="选择发票PDF文件（可多选，支持多页）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if files:
            file_list = list(files)
            self.create_progress_dialog(len(file_list))
            def import_thread():
                self.process_files(file_list)
                self.root.after(0, self.safe_close_loading_dialog)
            threading.Thread(target=import_thread, daemon=True).start()

    def process_files(self, files):
        added = 0
        failed = 0
        total = len(files)
        for idx, path in enumerate(files):
            self.root.after(0, self.update_progress, idx + 1, total)
            if path in [inv[0] for inv in self.invoices]:
                self.root.after(0, lambda p=path: self.tree.insert('', END, values=('☐', 0, os.path.basename(p), "已添加过")))
                self.root.after(0, self.renumber_treeview)
                failed += 1
                continue
            if not self.has_invoice_number_in_first_two_pages(path):
                self.root.after(0, lambda p=path: self.tree.insert('', END, values=('☐', 0, os.path.basename(p), "非发票")))
                self.root.after(0, self.renumber_treeview)
                failed += 1
                continue
            try:
                df = self.extract_invoice_with_precise_merge(path)
                if df is not None and not df.empty:
                    name = os.path.basename(path)
                    self.invoices.append((path, df, name))
                    rows_count = len(df)
                    self.root.after(0, lambda p=path, r=rows_count, n=name: (
                        self.tree.insert('', END, values=('✓', 0, n, r)) and
                        self.tree.item(self.tree.get_children()[-1], tags=('success',))
                    ))
                    self.root.after(0, self.renumber_treeview)
                    added += 1
                else:
                    self.root.after(0, lambda p=path: self.tree.insert('', END, values=('☐', 0, os.path.basename(p), "0")))
                    self.root.after(0, self.renumber_treeview)
                    failed += 1
            except Exception as e:
                self.root.after(0, lambda p=path: self.tree.insert('', END, values=('☐', 0, os.path.basename(p), "错误")))
                self.root.after(0, self.renumber_treeview)
                failed += 1
                traceback.print_exc()
        self.root.after(0, self.update_progress, total, total)
        self.root.after(0, self.renumber_treeview)
        self.root.after(0, self.update_selected_count)
        if added > 0 or failed > 0:
            self.root.after(0, lambda: self.show_result_dialog(added, failed))
        if failed == 0 and added > 0:
            self.root.after(0, lambda: self.all_var.set("1"))
            self.root.after(0, self.toggle_all)
        else:
            self.root.after(0, lambda: self.all_var.set("0"))

    def show_result_dialog(self, added, failed):
        dialog = Toplevel(self.root)
        dialog.title("处理结果")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog_width = 360
        dialog_height = 160
        self.root.update_idletasks()
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        pos_x = root_x + (root_width - dialog_width) // 2
        pos_y = root_y + (root_height - dialog_height) // 2
        dialog.geometry(f"{dialog_width}x{dialog_height}+{pos_x}+{pos_y}")
        frame = Frame(dialog, padx=20, pady=20)
        frame.pack(fill='both', expand=True)
        Label(frame, text=f"处理完成：成功 {added} 个，失败/未识别 {failed} 个\n\n"
                         f"成功识别的发票已自动勾选",
              font=('微软雅黑', 10), justify='left').pack()
        Button(frame, text="确定", width=10, command=dialog.destroy).pack(pady=10)

    def renumber_treeview(self):
        for i, item in enumerate(self.tree.get_children(), 1):
            values = list(self.tree.item(item, 'values'))
            values[1] = i
            self.tree.item(item, values=values)

    def delete_selected(self):
        selected = [item for item in self.tree.get_children() if self.tree.item(item, 'values')[0] == '✓']
        if not selected:
            messagebox.showinfo("提示", "未选中任何发票")
            return
        if not messagebox.askyesno("确认删除", f"确定要删除选中的 {len(selected)} 张发票吗？"):
            return
        for item in selected:
            self.tree.delete(item)
        remaining_filenames = [self.tree.item(item, 'values')[2] for item in self.tree.get_children()]
        self.invoices = [inv for inv in self.invoices if inv[2] in remaining_filenames]
        self.renumber_treeview()
        self.update_all_var()
        self.update_selected_count()
        messagebox.showinfo("删除完成", f"已删除 {len(selected)} 张发票")

    def export_to_excel(self):
        selected_items = [item for item in self.tree.get_children() if self.tree.item(item, 'values')[0] == '✓']
        if not selected_items:
            messagebox.showwarning("未选择", "请至少勾选一张发票进行导出")
            return

        selected_invoices = []
        for item in selected_items:
            display_name = self.tree.item(item, 'values')[2]
            for inv in self.invoices:
                if inv[2] == display_name:
                    selected_invoices.append((inv[0], inv[1].copy(), inv[2]))
                    break

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"发票明细_{ts}.xlsx"
        full_path = os.path.join(self.save_path, filename)

        try:
            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                if self.export_mode.get() == "merge":
                    dfs_to_merge = []
                    for _, df, _ in selected_invoices:
                        if '序号' in df.columns:
                            df = df.drop(columns=['序号'])
                        dfs_to_merge.append(df)
                    combined = pd.concat(dfs_to_merge, ignore_index=True)
                    combined.insert(0, '序号', range(1, len(combined) + 1))
                    combined.to_excel(writer, sheet_name="发票明细", index=False)
                    worksheet = writer.sheets['发票明细']
                    self.format_numeric_columns(worksheet, combined)
                    self.add_tax_formula(worksheet, combined)
                    total_row_idx = len(combined) + 2
                    worksheet.cell(row=total_row_idx, column=1).value = "合计"
                    amount_col_idx = None
                    tax_amount_col_idx = None
                    for col_idx, col_name in enumerate(combined.columns, 1):
                        if col_name == '金额':
                            amount_col_idx = col_idx
                        elif col_name == '税额':
                            tax_amount_col_idx = col_idx
                    if amount_col_idx:
                        letter = get_column_letter(amount_col_idx)
                        worksheet.cell(row=total_row_idx, column=amount_col_idx).value = f"=SUM({letter}2:{letter}{total_row_idx-1})"
                        worksheet.cell(row=total_row_idx, column=amount_col_idx).number_format = '0.00'
                    if tax_amount_col_idx:
                        letter = get_column_letter(tax_amount_col_idx)
                        worksheet.cell(row=total_row_idx, column=tax_amount_col_idx).value = f"=SUM({letter}2:{letter}{total_row_idx-1})"
                        worksheet.cell(row=total_row_idx, column=tax_amount_col_idx).number_format = '0.00'
                    for col in range(1, len(combined.columns) + 1):
                        cell = worksheet.cell(row=total_row_idx, column=col)
                        cell.font = Font(bold=True)
                else:
                    used_names = set()
                    for idx, (_, df, name) in enumerate(selected_invoices, 1):
                        if '序号' in df.columns:
                            df = df.drop(columns=['序号'])
                        df.insert(0, '序号', range(1, len(df) + 1))
                        base_name = f"发票{idx}_{os.path.splitext(name)[0][:20]}"
                        sheet_name = base_name[:31]
                        i = 1
                        while sheet_name in used_names:
                            sheet_name = f"{base_name[:28]}_{i}"[:31]
                            i += 1
                        used_names.add(sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        worksheet = writer.sheets[sheet_name]
                        self.format_numeric_columns(worksheet, df)
                        self.add_tax_formula(worksheet, df)
                        total_row_idx = len(df) + 2
                        worksheet.cell(row=total_row_idx, column=1).value = "合计"
                        amount_col_idx = None
                        tax_amount_col_idx = None
                        for col_idx, col_name in enumerate(df.columns, 1):
                            if col_name == '金额':
                                amount_col_idx = col_idx
                            elif col_name == '税额':
                                tax_amount_col_idx = col_idx
                        if amount_col_idx:
                            letter = get_column_letter(amount_col_idx)
                            worksheet.cell(row=total_row_idx, column=amount_col_idx).value = f"=SUM({letter}2:{letter}{total_row_idx-1})"
                            worksheet.cell(row=total_row_idx, column=amount_col_idx).number_format = '0.00'
                        if tax_amount_col_idx:
                            letter = get_column_letter(tax_amount_col_idx)
                            worksheet.cell(row=total_row_idx, column=tax_amount_col_idx).value = f"=SUM({letter}2:{letter}{total_row_idx-1})"
                            worksheet.cell(row=total_row_idx, column=tax_amount_col_idx).number_format = '0.00'
                        for col in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=total_row_idx, column=col)
                            cell.font = Font(bold=True)

            # 导出成功弹窗
            dialog = Toplevel(self.root)
            dialog.title("导出成功！")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.resizable(False, False)
            dialog_width = 420
            dialog_height = 180
            self.root.update_idletasks()
            root_x = self.root.winfo_x()
            root_y = self.root.winfo_y()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            pos_x = root_x + (root_width - dialog_width) // 2
            pos_y = root_y + (root_height - dialog_height) // 2
            dialog.geometry(f"{dialog_width}x{dialog_height}+{pos_x}+{pos_y}")

            msg_frame = Frame(dialog, padx=30, pady=20)
            msg_frame.pack(fill='both', expand=True)

            Label(msg_frame,
                  text=f"成功导出 {len(selected_invoices)} 张发票明细\n\n"
                       f"文件：{filename}\n"
                       f"路径：{os.path.normpath(self.save_path)}",
                  font=('微软雅黑', 10), justify='left').pack(anchor='w')

            btn_frame = Frame(dialog)
            btn_frame.pack(pady=(10, 20))

            def close_dialog():
                dialog.destroy()

            # 打开所在文件夹（立即关闭）
            Button(btn_frame, text="打开所在文件夹",
                   font=('微软雅黑', 10),
                   width=14,
                   command=lambda: [self.open_save_folder(), close_dialog()]).pack(side='left', padx=8)

            # 打开文件：异步打开，点击后立即关闭窗口
            def open_file_async():
                def target():
                    try:
                        if os.name == 'nt':
                            os.startfile(full_path)
                        else:
                            subprocess.call(['xdg-open', full_path])
                    except Exception:
                        pass
                threading.Thread(target=target, daemon=True).start()

            Button(btn_frame, text="打开文件",
                   font=('微软雅黑', 10),
                   width=10,
                   command=lambda: [open_file_async(), close_dialog()]).pack(side='left', padx=8)

            # 关闭
            Button(btn_frame, text="关闭",
                   font=('微软雅黑', 10),
                   width=10,
                   command=close_dialog).pack(side='left', padx=8)

        except Exception as e:
            messagebox.showerror("导出失败", f"导出过程中发生错误：\n{str(e)}")
            traceback.print_exc()

    def format_numeric_columns(self, worksheet, df):
        numeric_cols = ['数量', '单价', '金额']
        col_indices = {}
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in numeric_cols:
                col_indices[col_name] = col_idx

        for col_name, col_idx in col_indices.items():
            letter = get_column_letter(col_idx)
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    original_str = str(df.iloc[row_idx-2][col_name])
                    if original_str and original_str != '':
                        clean_str = original_str.replace(',', '')
                        num_val = float(clean_str)
                        cell.value = num_val
                        if '.' in original_str:
                            decimals = len(original_str.split('.')[-1])
                            fmt = f'0.{"0" * decimals}'
                        else:
                            fmt = '0'
                        cell.number_format = fmt
                except:
                    cell.number_format = '0.00'

    def add_tax_formula(self, worksheet, df):
        amount_col_idx = tax_rate_col_idx = tax_amount_col_idx = None
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == '金额':
                amount_col_idx = col_idx
            elif col_name == '税率/征收率':
                tax_rate_col_idx = col_idx
            elif col_name == '税额':
                tax_amount_col_idx = col_idx

        if amount_col_idx and tax_rate_col_idx and tax_amount_col_idx:
            for row_idx in range(2, len(df) + 2):
                amount_cell = f"{get_column_letter(amount_col_idx)}{row_idx}"
                tax_rate_cell = f"{get_column_letter(tax_rate_col_idx)}{row_idx}"
                tax_amount_cell = f"{get_column_letter(tax_amount_col_idx)}{row_idx}"
                formula = f'=TRUNC({amount_cell}*SUBSTITUTE({tax_rate_cell},"%","")/100, 2)'
                worksheet[tax_amount_cell] = formula

    def extract_invoice_with_precise_merge(self, pdf_path):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                all_goods_lines = []
                for page_num, page in enumerate(pdf.pages, 1):
                    full_text = page.extract_text()
                    if not full_text:
                        continue
                    lines = [line.strip() for line in full_text.split('\n') if line.strip()]
                    goods_section = self.find_goods_section_smart(lines)
                    if goods_section:
                        all_goods_lines.extend(goods_section)

                if not all_goods_lines:
                    return None

                parsed_data = []
                for line in all_goods_lines:
                    if self.is_total_line(line):
                        continue
                    row_data = self.parse_goods_line_corrected(line)
                    parsed_data.append(row_data)

                if not parsed_data:
                    return None

                df = pd.DataFrame(parsed_data)
                required_cols = ['项目名称', '规格型号', '单位', '数量', '单价', '金额', '税率/征收率', '税额']
                for col in required_cols:
                    if col not in df.columns:
                        df[col] = ""

                df = df[required_cols]
                df = self.merge_continued_rows(df)
                df['项目名称'] = df['项目名称'].apply(self.clean_chinese_text)
                df['规格型号'] = df['规格型号'].apply(self.clean_chinese_text)
                df.insert(0, '序号', range(1, len(df) + 1))
                return df
        except Exception as e:
            traceback.print_exc()
            return None

    def parse_goods_line_corrected(self, line):
        row_data = {
            '项目名称': '', '规格型号': '', '单位': '',
            '数量': '', '单价': '', '金额': '',
            '税率/征收率': '', '税额': ''
        }

        original_line = line
        line = line.strip()
        if not line:
            return row_data

        tax_rate_match = re.search(r'(\d+(?:\.\d+)?%)', line)
        if tax_rate_match:
            row_data['税率/征收率'] = tax_rate_match.group(1)

        unit = None
        unit_pos = -1
        for u in UNIT_KEYWORDS:
            pos = line.find(u)
            if pos != -1:
                context_start = max(0, pos-5)
                context_end = min(len(line), pos+len(u)+5)
                context = line[context_start:context_end]
                if re.search(r'\d', context):
                    unit = u
                    unit_pos = pos
                    break

        if unit:
            row_data['单位'] = unit

        if unit and unit_pos != -1:
            right_part = line[unit_pos + len(unit):].replace(',', '')
            numbers_raw = re.findall(r'\d+(?:\.\d+)?', right_part)
        else:
            line = line.replace(',', '')
            numbers_raw = re.findall(r'\d+(?:\.\d+)?', line)

        numbers = numbers_raw

        tax_rate_num = row_data['税率/征收率'].rstrip('%') if row_data['税率/征收率'] else None
        if tax_rate_num and tax_rate_num in numbers:
            numbers = numbers[::-1]
            numbers.remove(tax_rate_num)
            numbers = numbers[::-1]

        if len(numbers) >= 4:
            row_data['数量'] = numbers[0]
            row_data['单价'] = numbers[1]
            row_data['金额'] = numbers[2]
            row_data['税额'] = numbers[3]
        elif len(numbers) == 3:
            row_data['数量'] = numbers[0]
            row_data['单价'] = numbers[1]
            row_data['金额'] = numbers[2]
        elif len(numbers) == 2:
            row_data['数量'] = numbers[0]
            row_data['单价'] = numbers[1]
        elif len(numbers) == 1:
            row_data['金额'] = numbers[0]

        left_text = original_line
        if unit and unit_pos != -1:
            left_text = original_line[:unit_pos].strip()
        else:
            if numbers_raw:
                for num in numbers_raw:
                    pos = original_line.find(num)
                    if pos != -1:
                        left_text = original_line[:pos].strip()
                        break

        left_text = left_text.replace('**', '')
        if ' ' in left_text:
            parts = left_text.split(' ')
            if len(parts) >= 2:
                row_data['项目名称'] = parts[0].strip()
                row_data['规格型号'] = ' '.join(parts[1:]).strip()
            else:
                row_data['项目名称'] = left_text.strip()
        else:
            row_data['项目名称'] = left_text.strip()

        return row_data

    def merge_continued_rows(self, df):
        if df.empty:
            return df

        key_cols = ['单位', '数量', '单价', '金额']
        has_key = df[key_cols].notna() & (df[key_cols] != '')
        is_complete = has_key.any(axis=1)

        merged_rows = []
        current_row = None

        for idx, row in df.iterrows():
            name_text = str(row['项目名称']).strip()
            model_text = str(row['规格型号']).strip()

            if is_complete[idx]:
                if current_row is not None:
                    merged_rows.append(current_row)
                current_row = row.copy()
            else:
                if current_row is None:
                    current_row = row.copy()
                else:
                    # 续行内容追加到规格型号（无论在name_text还是model_text）
                    append_text = (name_text + " " + model_text).strip()
                    if append_text:
                        if current_row['规格型号']:
                            current_row['规格型号'] += " " + append_text
                        else:
                            current_row['规格型号'] = append_text

        if current_row is not None:
            merged_rows.append(current_row)

        return pd.DataFrame(merged_rows).reset_index(drop=True)

    def clean_chinese_text(self, text):
        if not text:
            return ""
        cleaned = re.sub(r'\s+', ' ', str(text)).strip()
        cleaned = re.sub(r'\*\s+', '*', cleaned)
        cleaned = re.sub(r'\s+\*', '*', cleaned)
        cleaned = re.sub(r'\(\s+', '(', cleaned)
        cleaned = re.sub(r'\s+\)', ')', cleaned)
        cleaned = re.sub(r'([\u4e00-\u9fff])\s+([\u4e00-\u9fff])', r'\1\2', cleaned)
        return cleaned

    def find_goods_section_smart(self, lines):
        goods_section = []
        in_goods_section = False

        for line in lines:
            if not in_goods_section:
                if self.is_goods_header(line):
                    in_goods_section = True
                    continue

            if in_goods_section:
                if self.is_end_of_goods_section(line):
                    break
                goods_section.append(line)

        return goods_section

    def is_goods_header(self, line):
        keywords = ['项目名称', '规格型号', '单位', '数量', '单价', '金额', '税率']
        return sum(1 for kw in keywords if kw in line) >= 2

    def is_end_of_goods_section(self, line):
        if any(kw in line for kw in ['合计', '价税合计', '小计', '总计']):
            if re.search(r'\d{3,}', line.replace(',', '')):
                return True

        if any(m in line for m in ['税额', '¥', '￥', '备注', '开票人', '收款人']) and re.search(r'\d', line):
            return True

        return False

    def is_total_line(self, line):
        if any(kw in line for kw in ['合计', '价税合计', '小计', '总计']):
            if re.search(r'\d{3,}', line.replace(',', '')):
                return True
        return False

    def set_save_path(self):
        p = filedialog.askdirectory(title="选择导出文件夹", initialdir=self.save_path)
        if p:
            self.save_path = os.path.normpath(p)
            self.path_label.config(text=self.shorten_path(self.save_path))
            messagebox.showinfo("成功", f"保存路径已更新为：\n{self.save_path}")

    def open_save_folder(self):
        try:
            if os.path.exists(self.save_path):
                if os.name == 'nt':
                    os.startfile(self.save_path)
                else:
                    subprocess.call(['xdg-open', self.save_path])
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件夹：\n{str(e)}")


if __name__ == "__main__":
    root = Tk()
    app = InvoiceApp(root)
    root.mainloop()
